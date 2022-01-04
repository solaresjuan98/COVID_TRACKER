import base64
import datetime as dt
import io
from posixpath import abspath
import time
from math import e

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import plotly
import plotly.express as px
import plotly.figure_factory as ff
import seaborn as sns
import streamlit as st
from attr import field
from fpdf import FPDF
from matplotlib import colors
from pandas._config.config import options
from pandas.core import groupby
from pandas.core.algorithms import mode
from pandas.core.frame import DataFrame
from pandas.core.indexes.timedeltas import timedelta_range
from pandas.core.reshape.pivot import pivot_table
from PIL import Image
from sklearn import linear_model
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.preprocessing import PolynomialFeatures
from streamlit.elements.arrow import Data

prediccion = 0


def generatePredictionGraph(y: DataFrame, grade, days, max_val):

    X = []
    Y = y
    print(y)

    size = y.__len__()
    for i in range(0, size):
        X.append(i)

    #st.write('XD')
    #st.write(Y)
    X = np.asarray(X)
    Y = np.asarray(Y)

    X = X[:, np.newaxis]
    Y = Y[:, np.newaxis]

    # GRAPH 1
    #st.subheader("Plot graph")
    #st.set_option('deprecation.showPyplotGlobalUse', False)
    #plt.scatter(X, Y)
    #plt.show()
    #st.pyplot()

    # st.write(Y)
    #  2:
    nb_degree = grade

    polynomial_features = PolynomialFeatures(degree=nb_degree)
    X_TRANSF = polynomial_features.fit_transform(X)

    # ## print(Y)
    #  3:
    model = LinearRegression()
    model.fit(X_TRANSF, Y)

    #  4
    Y_NEW = model.predict(X_TRANSF)

    rmse = np.sqrt(mean_squared_error(Y, Y_NEW))
    r2 = r2_score(Y, Y_NEW)

    print('RMSE: ', rmse)
    print('R2: ', r2)

    #  5:
    x_new_min = 0.0
    x_new_max = float(days)  ## days to predict

    X_NEW = np.linspace(x_new_min, x_new_max, 50)
    X_NEW = X_NEW[:, np.newaxis]

    X_NEW_TRANSF = polynomial_features.fit_transform(X_NEW)

    Y_NEW = model.predict(X_NEW_TRANSF)

    st.subheader("Plot graph")
    plt.scatter(X, Y)
    plt.plot(X_NEW, Y_NEW, color='green', linewidth=3)
    #plt.scatter(X_NEW, Y_NEW, color='blue', linewidth=3)
    plt.grid()
    plt.xlim(x_new_min, x_new_max)  ## X axis

    plt.ylim(0, Y_NEW[int(Y_NEW.size - 1)])
    title = 'Degree={}; RMSE={}; R2={}'.format(nb_degree, round(rmse, 2),
                                               round(r2, 2))
    plt.title(title)
    plt.xlabel('x')
    plt.ylabel('y')

    plt.savefig('prediction.png')
    #plt.savefig('pol_reg.jpg', )
    plt.show()
    st.pyplot()
    st.caption('Prediction graph')
    global prediccion
    prediccion = Y_NEW[int(Y_NEW.size - 1)][0]
    st.write("La predicción será de ", Y_NEW[int(Y_NEW.size - 1)][0])

    #
    #fig = ff.create_distplot(Y_NEW, ['test'], bin_size=[0] )

    pass


def generateTendencyGraph(y, header, maxY):
    x = []
    #y = country_infections[select_col[2]]

    for i in range(0, y.__len__()):
        x.append(i)

    X = np.asarray(x).reshape(-1, 1)
    reg = linear_model.LinearRegression()
    reg.fit(X, y)
    y_pred = reg.predict(X)

    plt.scatter(X, y, color='black')
    plt.plot(X, y_pred, color='blue', linewidth=3)
    plt.savefig('trend.jpg')
    #print(reg.predict([[10]]))
    max_val = maxY

    # Header
    st.subheader(header)
    plt.title(header)
    plt.ylim(-10, max_val + 10)
    plt.show()
    st.pyplot()
    st.set_option('deprecation.showPyplotGlobalUse', False)
    st.caption('COVID-19 tendence graph')

    # st.info("""Para poder comprender de mejor forma esta grafica,
    # es importante tomar en cuenta la pendiente generada, por ejemplo si la pendiente es creciente (positiva)
    # la tendencia de contagios en los dias siguientes al analisis será a la alta, pero si de lo contrario
    # la pendiente de la grafica es decreciente (negativa), la tendencia numero de contagios es a la baja.
    # """)


def create_download_link(val, filename):
    b64 = base64.b64encode(val)  # val looks like b'...'
    return f'<a href="data:application/octet-stream;base64,{b64.decode()}" download="{filename}.pdf">Download file</a>'


# write pdf report
def write_pdf(title, content, imagepath):

    pdf = FPDF()
    pdf.add_page()
    #pdf.set_xy(0, 0)
    pdf.set_font('Times', 'B', 12)
    pdf.cell(200, 10, title, 0, 2, 'C')
    pdf.set_font('Times', '', 12)
    pdf.multi_cell(200, 10, txt=content, align="J")

    #pdf.cell(90, 10, " ", 0, 2, 'C')
    #pdf.cell(-40)
    # pdf.cell(50, 10, 'X', 1, 0, 'C')
    # pdf.cell(40, 10, variable[0], 1, 0, 'C')
    # pdf.cell(40, 10, '', 1, 2, 'C')
    # pdf.cell(-90)
    # pdf.set_font('arial', '', 12)
    # for i in range(0, len(flt)):
    #     pdf.cell(50, 10, '%s' % (flt[variable[1]].iloc[i]),
    #              1, 0, 'C')
    #     pdf.cell(40, 10,
    #              '%s' % (str(flt[variable[0]].iloc[i])), 1,
    #              0, 'C')
    #     pdf.cell(40, 10,
    #              '%s' % (str(flt.positive.iloc[i])), 1, 2,
    #              'C')
    #     pdf.cell(-90)
    # pdf.cell(90, 10, " ", 0, 2, 'C')

    pdf.image(imagepath)
    html = create_download_link(pdf.output(dest="S").encode("latin-1"), "test")

    st.markdown(html, unsafe_allow_html=True)


# ===================== REPORTS =====================


# 1. Tendencia de la infección por Covid-19 en un País.
def covidInfectionTendence(data: DataFrame):

    data_options = st.multiselect('Select filter [country]: ', data.columns)
    # st.write(data_options)
    try:

        if data_options.__len__() == 0:
            st.warning('Please select a column')
        else:

            country_options = st.selectbox(
                'Select country', data[data_options[0]].drop_duplicates())

            country = [country_options]

            flt = data[data[data_options[0]].isin(country)]

            have_date = st.checkbox('This file has "date" field')

            #st.write(have_date)

            if have_date:
                variable = st.multiselect(
                    'Select variables to analize [date, numeric]: ',
                    data.columns)

                if variable.__len__() < 2:
                    st.warning('Please select variables to analize')

                else:

                    st.write(variable)
                    # data

                    flt[variable[0]] = pd.to_datetime(flt[variable[0]])
                    flt[variable[0]] = flt[variable[0]]

                    flt = flt[[variable[0],
                               variable[1]]].sort_values(by=[variable[0]])

                    #sum
                    st.write(
                        flt.groupby([variable[0],
                                     variable[1]]).sum().reset_index())

                    # Tendency
                    generateTendencyGraph(flt[variable[1]],
                                          "Tendency COVID spread by country",
                                          flt[variable[1]].max())

                    st.subheader('Intepretación: ')
                    st.info(
                        """Esta grafica muestra la tendencia de indice de contagios que tiene {} en un periodo de {} días. 
                        Una pendiente positiva indica que la cantidad de casos irán aumentando con el paso del tiempo, y una pendiente 
                        negativa indica que los casos irán decreciendo con el paso del tiempo. """
                        .format(country[0], flt.__len__()))

                    ### PDF
                    pdf_title = 'Tendencia de la infección por Covid-19 en un País. \n\n'
                    content = """Esta grafica muestra la tendencia de indice de contagios que tiene {} en un periodo de {} días. 
                        Una pendiente positiva indica que la cantidad de casos irán aumentando con el paso del tiempo, y una pendiente negativa indica que los casos irán decreciendo con el paso del tiempo. """.format(
                        country[0], flt.__len__())

                    export_as_pdf = st.button("Export Report")

                    if export_as_pdf:
                        write_pdf(pdf_title, content, 'D:\\trend.jpg')

            else:
                variable = st.selectbox('Select variable to analyze: ',
                                        data.columns)

                st.write(flt[variable])

                generateTendencyGraph(flt[variable],
                                      "Tendency COVID spread by country",
                                      flt[variable].max())

            #generatePredictionGraph(flt[variable], 3, 40, flt[variable].max())

    except Exception as e:
        st.write(e)
        st.warning("Please select a field")


# 2. Predicción de Infectados en un País.
def covidInfectedPredictionByCountry(data: DataFrame):

    has_date = st.checkbox('This file has "date" field')

    try:
        if has_date:
            option = st.multiselect(
                'Select date, country and numeric variable: ', data.columns)

            if option.__len__() < 3:
                st.warning('Please select more fields')
            else:

                df = data[[option[0], option[1], option[2]]]
                #st.write(df)
                country = st.selectbox('Select country: ',
                                       df[[option[1]]].drop_duplicates())

                # Filter data by country
                c = [country]

                data[data[option[1]].isin(c)]

                y = data[option[2]]

                days = st.slider('Select a number of days to predict', 5, 1000)

                grade = st.slider('Select a polynomial grade prediction: ', 1,
                                  5)
                # print(y)
                generatePredictionGraph(y, grade, days, y.max())

                interpretacion = """
                La gráfica muestra una predicción de personas infectadas en {} en un periodo de {} días, y la predicción está siendo representada por medio de un polinomio
                de grado {}. (Es importante tomar en cuenta que entre mayor sea el grado de un polinomio mucho más precisa será la predicción obtenida).
                """.format(c, days, grade)

                st.info(interpretacion)

                export_as_pdf = st.button("Export Report")
                pdf_title = '2.  Predicción de Infectados en un País.'
                #content = ""

                if export_as_pdf:
                    write_pdf(pdf_title, interpretacion, 'prediction.png')
                pass

        else:

            option = st.multiselect(
                'Select a field, and numeric variable to filter [ex. country, infections]: ',
                data.columns)

            country = st.selectbox('Select country',
                                   data[option[0]].drop_duplicates())

            data = data[data[option[0]].isin([country])]

            data[[option[0], option[1]]]

            days = st.slider('Select a number of days to predict', 5, 1000)

            grade = st.slider('Select a polynomial grade prediction: ', 1, 5)

            y = data[option[1]]
            generatePredictionGraph(y, grade, days, y.max())
            #st.write(data)
            export_as_pdf = st.button("Export Report")
            pdf_title = '2.  Predicción de Infectados en un País.'
            content = ""

            if export_as_pdf:
                write_pdf(pdf_title, content, 'prediction.png')
            pass

    except Exception as e:
        st.write(e)
        st.warning('Please select three fields')

    # infected = st.multiselect('Select numeric variable: ', data.columns)

    pass


# 3. Indice de Progresión de la pandemia.
def pandemicProgression(data: DataFrame):

    options = st.multiselect('Select variable and order by [cases, date]: ',
                             data.columns)

    if options.__len__() < 2:

        st.warning('Please select a numeric variable and date')

    else:

        cases = options[0]
        date_ = options[1]

        data = data[[date_, cases]]

        # data['fecha'] = pd.to_datetime(data['fecha'])
        data[date_] = pd.to_datetime(data[date_])
        # data = data.sort_values(by='fecha')
        data = data.sort_values(by=date_)
        data = data.groupby(date_)[cases].sum()
        st.write(data)

        st.subheader('Global pandemic progression')

        st.line_chart(data)
        st.caption('Pandemic progression')
        st.info("""
        Esta grafica representa el progreso de infecciones que ha habido desde el inicio de la pandemia en todo el mundo
        """)

        generateTendencyGraph(data, 'Pandemic progression regression',
                              data.max())

        interpretacion = """
        La grafica de tendencia representa el comportamiento que tendrá la variable estudiada "{}" a lo largo del tiempo, y la pendiente indica 
        si la tendencia irá aumentando o disminuyendo con el tiempo.
        """.format(cases)

        st.info(interpretacion)

        export_as_pdf = st.button("Export Report")
        pdf_title = '3. Indice de Progresión de la pandemia.'
        #content = ""

        if export_as_pdf:
            write_pdf(pdf_title, interpretacion, 'trend.jpg')


# 4. Predicción de mortalidad por COVID en un Departamento.
def covidDeathsPredictionByDeparment(data: DataFrame):

    try:
        # date, state, cases
        data_options = st.multiselect(
            'Select fields [date, region, cases, filter]: ', data.columns)

        if data_options.__len__() < 4:
            st.warning('Select more fields to analyze')

        else:

            date_ = data_options[0]
            region = data_options[1]
            cases = data_options[2]
            flter = data_options[3]

            st.write(data_options)

            country_option = st.selectbox('Select country',
                                          data[region].drop_duplicates())

            ## select states
            c = [country_option]
            cs = data[data[region].isin(c)]

            province = st.selectbox('Select state/province/department',
                                    cs[flter].drop_duplicates())

            # Filter by deparment/state
            dep = cs[data[flter].isin([province])]

            ## convert dates
            dep[date_] = pd.to_datetime(dep[date_])
            dep = dep.sort_values(by=date_)
            st.subheader('Deaths in {} by day'.format(province))
            st.write(dep[[date_, cases]])

            y = dep[cases]

            # slider
            n_days = st.slider('Select number of days to predict: ', 5, 100)
            grade = st.slider('Select grade of the regression: ', 1, 3)
            max_val = dep[cases].max()

            generatePredictionGraph(y, grade, n_days, max_val)

            interpretacion = """
            La gráfica de predicción nos indica el posible valor de la variable "{}" en un periodo de {} días, 
            representado con un polinomio de grado {}.
            """.format(cases, n_days, grade)

            st.info(interpretacion)

            export_as_pdf = st.button("Export Report")
            pdf_title = '4. Predicción de mortalidad por COVID en un Departamento.'
            #content = ""

            if export_as_pdf:
                write_pdf(pdf_title, interpretacion, 'prediction.png')

    except Exception as e:
        st.write(e)
        st.warning('Select a field')


# 5. Predicción de mortalidad por COVID en un País
def covidDeathPredictionByCountry(data: DataFrame):

    try:

        options_selected = st.multiselect(
            'Select fields to filter [country, deaths]: ', data.columns)
        st.write(options_selected)
        if options_selected.__len__() < 2:
            st.warning('Plase select fields to filter')
        else:
            fltr = options_selected[1]
            country = st.selectbox('Please select a country: ',
                                   data[options_selected[0]].drop_duplicates())

            data = data[data[options_selected[0]].isin([country])]
            data[options_selected[0]] = data[options_selected[0]].fillna(0)
            has_date = st.checkbox('The file has "date" field')
            #data.fillna(0)

            if has_date:

                date_ = st.selectbox('Select a type "date" field ',
                                     data.columns)

                data[date_] = pd.to_datetime(data[date_])

                # flt = data[[date_, fltr]]
                # st.write(flt.reset_index())
                # flt = flt.sort_values(by=date_)
                # st.write(flt)
                data = data.sort_values(by=date_).reset_index()

                #data.fillna(0)

                #data[['dateRep', 'deaths']]
                data[fltr] = data[fltr].fillna(0)

                n_days = st.slider('Select a number of days to predict: ', 100,
                                   1000)
                grade = st.slider('Select grade of the polynomial grade: ', 1,
                                  10)

                generatePredictionGraph(data[fltr], grade, n_days,
                                        data[fltr].max())

                export_as_pdf = st.button("Export Report")
                pdf_title = '5. Predicción de mortalidad por COVID en un País '

                global prediccion
                interpretacion = """
                La prediccion de numero de muertes en globalmente en un numero de {} días representado con un polnomio de  grado {}
                sera de {} personas, de acuerdo al dataset que se tiene.
                """.format(n_days, grade, int(prediccion))

                if export_as_pdf:
                    write_pdf(pdf_title, interpretacion, 'prediction.png')
            else:
                pass

    except Exception as e:
        st.write(e)
        st.warning('An error has occurred')

        #

        #flt[[]]

        #st.write(options_selected)


# 6. Análisis del número de muertes por coronavirus en un País.
def covidDeathsByCountry(data: DataFrame):

    data_options = st.multiselect(
        'Select field, and variables to analize [ex. country, deaths, date]: ',
        data.columns)

    try:

        if data_options.__len__() < 3:
            st.write('Please select more fields')
        else:

            # IMPORTANTE REVISAR AQUI XD
            country = st.selectbox('Select country',
                                   data[data_options[0]].drop_duplicates())

            data = data[data[data_options[0]].isin([country])]

            data[data_options[2]] = pd.to_datetime(data[data_options[2]])

            st.subheader('Deaths analysis in {}'.format(country))
            st.write(data[[data_options[1], data_options[2]]])

            y = data[data_options[1]].fillna(0)

            generateTendencyGraph(y, 'Deaths in {}'.format(country), y.max())
            intepretacion = """
            Esta grafica analiza el comportamiento del numero de muertes en {} desde que empezó la pandemia, así como la pendiente que indicará si el numero
            de fallecidos a causa del COVID-19 va a aumentar o disminuir con el paso del tiempo. Si la pendiente es creciente, indica que la cantidad de muertes
            tiende a crecer, pero si la pendiente es decreciente el numero de muertes tiende a disminuir con el tiempo.
            """.format(country)
            st.info(intepretacion)

            export_as_pdf = st.button("Export Report")
            pdf_title = '6. Análisis del número de muertes por coronavirus en un País.'
            #content = ""

            if export_as_pdf:
                write_pdf(pdf_title, intepretacion, 'trend.jpg')

    except Exception as e:
        st.write(e)
        st.warning("Please select a field")


# 7. Tendencia del número de infectados por día de un País
def covidInfectedByDay(data: DataFrame):

    select_col = st.multiselect(
        'Select a columns to parameterize (date, country and numeric variable): ',
        data.columns)

    #st.write(select_col)

    try:
        # select_date = st.selectbox('Select a date: ',
        #                            data[select_col[0]].drop_duplicates())

        if select_col.__len__() < 3:
            st.warning('Select more fields')
        else:
            select_country = st.selectbox(
                'Select a country: ', data[select_col[1]].drop_duplicates())

            ## Filter data
            country = [select_country]

            country_infections = data[data[select_col[1]].isin(country)]  ##

            ## sort
            country_infections[select_col[0]] = pd.to_datetime(
                country_infections[
                    select_col[0]])  #country_infections.sort_values(by='Date')
            st.write(country_infections)

            # group by and sum
            #st.write(country_infections.groupby([select_col[0], select_col[1]]).sum().sort_values(by='date'))

            fig = country_infections.groupby(
                [select_col[0],
                 select_col[1]]).sum().sort_values(by=select_col[0]).head()

            st.write(fig)

            # x = []
            y = country_infections[select_col[2]].fillna(0)
            hd = "COVID-19 Spread tendency in " + country[0]
            max_val = country_infections[select_col[2]].max()

            generateTendencyGraph(y, hd, max_val)
            interpretacion = """
            La gráfica muestra la tendencia del número diario de infectados en {}, al observar detenidamente la grafica, si la pendiente es creciente
            esto significa que el numero de contagios en {} estará aumentando con el paso del tiempo, pero si la pendiente es decreciente significa que el numero
            de contagios diarios irá decreciendo con el paso del tiempo, o por lo menos hasta la llegada de una nueva ola de contagios.
            """.format(select_country, select_country)

            st.info(interpretacion)

            export_as_pdf = st.button("Export Report")
            pdf_title = '7. Tendencia del número de infectados por día de un País'
            #content = ""

            if export_as_pdf:
                write_pdf(pdf_title, interpretacion, 'D:\\trend.jpg')

    except Exception as e:
        st.write(e)
        st.warning('Select the fields ')

    pass


# 8. Predicción de casos de un país para un año.
def casesPredictionOneYear(data: DataFrame):

    select_option = st.selectbox('Select a field to filter [country]: ',
                                 data.columns)

    country = st.selectbox('Select a country: ',
                           data[select_option].drop_duplicates())

    data = data[data[select_option].isin([country])]

    var = st.selectbox('Select a variable to predict', data.columns)

    has_date = st.checkbox('This file has "date" field')

    if has_date:

        date_ = st.selectbox('Order by: ', data.columns)

        data[date_] = pd.to_datetime(data[date_])

        flt = data[[date_, var]]
        st.write(flt)
        flt = data.sort_values(by=date_)

        st.write(flt[[date_, var]])

        grade = st.slider('Select polynomial grade: ', 1, 10)
        generatePredictionGraph(flt[var].fillna(0), grade, 365,
                                flt[var].fillna(0).max())
        global prediccion
        interpretacion = """
        La gráfica nos indica la predicción de casos que tendrá {} despues de un año del primer contagio por COVID-19, es importante tomar en cuenta
        que es un valor estimado, y que el grado del polinomio usado tambien indica la precisión de los valores obtenidos.
        La prediccion de casos para un año a partir del dia del primer caso detectado será de {} casos.
        """.format(country, int(prediccion))
        st.info(interpretacion)

        export_as_pdf = st.button("Export Report")
        pdf_title = '8. Predicción de casos de un país para un año.'
        #content = ""

        if export_as_pdf:
            write_pdf(pdf_title, interpretacion, 'prediction.png')

    else:

        y = data[var]

        days = st.slider('Select number of days to predict: ', 10, 1000)
        grade = st.slider('Select polynomial grade: ', 1, 10)

        st.write(y)
        generatePredictionGraph(y, grade, days, y.max())
        interpretacion = """
        La gráfica nos indica la predicción de casos que tendrá {} despues de un año del primer contagio por COVID-19, es importante tomar en cuenta
        que es un valor estimado, y que el grado del polinomio usado tambien indica la precisión de los valores obtenidos.
        """.format(country)
        st.write(interpretacion)

        export_as_pdf = st.button("Export Report")
        pdf_title = '8. Predicción de casos de un país para un año.'
        #content = ""

        if export_as_pdf:
            write_pdf(pdf_title, interpretacion, 'D:\\prediction.png')

    #generatePredictionGraph(data[var],6, 365, 500000)


# 9. Tendencia de la vacunación de en un País.
def vaccinationTendencyByCountry(data: DataFrame):

    data_options = st.multiselect(
        'Select [date, country/region, and numeric variable]: ', data.columns)

    try:

        if data_options.__len__() < 3:
            st.warning('Please select fields to filter')
        else:
            region = data_options[1]

            country = st.selectbox('Select country: ',
                                   data[region].drop_duplicates())

            # pais = 'Guatemala'
            # data = data[data[region].isin(['Guatemala'])]

            data = data[data[region].isin([country])]

            st.write(data)

            vac = data[data_options[2]].fillna(0)

            st.write(vac)

            generateTendencyGraph(vac, "Vaccines trend graph", vac.max() + 500)

            interpretacion = """Para poder comprender de mejor forma esta grafica,
            es importante tomar en cuenta la pendiente generada, por ejemplo si la pendiente es creciente (positiva)
            la tendencia de vacunacion COVID-19 en los dias siguientes al analisis será a la alta, pero si de lo contrario
            la pendiente de la grafica es decreciente (negativa), la tendencia numero de vacunaciones por COVID-19 es a la baja.
            """

            st.info(interpretacion)

            export_as_pdf = st.button("Export Report")
            pdf_title = '9. Tendencia de la vacunación de en un País.'
            #content = ""

            if export_as_pdf:
                write_pdf(pdf_title, interpretacion, 'trend.jpg')

    except Exception as e:
        st.write(e)
        st.warning(":c")


# 10. Ánalisis Comparativo de Vacunación entre 2 paises.
def vaccinationComparationByCountries(data: DataFrame):

    try:

        option = st.multiselect(
            'Select field and variable of comparation [place, variable, group by] : ',
            data.columns)

        if option.__len__() < 3:
            st.warning('Select more fields')

        else:

            place = option[0]
            variable = option[1]
            col1, col2 = st.columns(2)

            with col1:

                st.subheader('Select Place 1:')
                country1 = st.selectbox('Select country 1: ',
                                        data[place].drop_duplicates())
                p1 = data[data[place].isin([country1])]

                #st.write(p1[variable])
                t1 = p1[variable].sum()
                generateTendencyGraph(p1[variable],
                                      'Vaccinations in {}'.format(country1),
                                      p1[variable].fillna(0).max())

            with col2:

                st.subheader('Select Place 2:')
                country2 = st.selectbox('Select country 2: ',
                                        data[place].drop_duplicates())
                p2 = data[data[place].isin([country2])]

                #st.write(p2[variable])
                t2 = p2[variable].sum()
                generateTendencyGraph(p2[variable],
                                      'Vaccinations in {}'.format(country2),
                                      p2[variable].fillna(0).max())

            ## graph
            plotdata = pd.DataFrame({str(variable): [t1, t2]},
                                    index=[country1, country2])

            st.bar_chart(plotdata)
            st.caption('Vaccination comparative graph')

            ## Interpretación
            interpretacion = ""
            if t1 > t2:
                st.info(
                    'De acuerdo a la grafica, {} presenta una mejor tasa de vacunacion que {}'
                    .format(country1, country2))
                interpretacion += 'De acuerdo a la grafica, {} presenta una mejor tasa de vacunacion que {}'.format(
                    country1, country2)

            else:

                st.info(
                    'De acuerdo a la grafica, {} presenta una mejor tasa de vacunacion que {}'
                    .format(country1, country2))
                interpretacion += 'De acuerdo a la grafica, {} presenta una mejor tasa de vacunacion que {}'.format(
                    country1, country2)

            st.info(
                """Es importante analizar los datos de las vacunaciones, ya que en este aspecto podemos analizar las diferentes situaciones
            en los que puede encontrar cada pais, los países más desarrollados tienen acceso más rapido a las vacunas que los paises menos desarrollados
            debido a factores como el economico, y politico (como la corrupción) o incluso la cultura de los países, ya que existen grupos
            de personas que debido a creencias culturales, religiosas, etc. se rehusan a vacunarse."""
            )

            interpretacion += """Es importante analizar los datos de las vacunaciones, ya que en este aspecto podemos analizar las diferentes situaciones
            en los que puede encontrar cada pais, los países más desarrollados tienen acceso más rapido a las vacunas que los paises menos desarrollados
            debido a factores como el economico, y politico (como la corrupción) o incluso la cultura de los países, ya que existen grupos
            de personas que debido a creencias culturales, religiosas, etc. se rehusan a vacunarse."""

            export_as_pdf = st.button("Export Report")
            pdf_title = '10. Ánalisis Comparativo de Vacunación entre 2 paises.'
            #content = ""

            if export_as_pdf:
                write_pdf(pdf_title, interpretacion, 'trend.jpg')

            # st.set_option('deprecation.showPyplotGlobalUse', False)
            # plotdata.plot(kind="bar", color="green")
            # plt.title("Comparative")
            # st.pyplot()

    except Exception as e:

        st.write(e)
        st.warning('Error :(')


# 11. Porcentaje de hombres infectados por covid-19 en un País desde el primer caso activo
def menPercentageInfected(data: DataFrame):

    field = st.selectbox('Select field to filter: ', data.columns)

    try:

        country = st.selectbox('Select country: ',
                               data[field].drop_duplicates())

        data = data[data[field].isin([country])]

        gender = st.multiselect('Select variables to analize [men, women]: ',
                                data.columns)

        if gender.__len__() < 2:
            st.warning('Select fields to compare. ')

        else:

            gender1 = gender[0]
            gender2 = gender[1]

            #data[[gender1, gender2]]
            t1 = data[gender1].fillna(0).sum()
            t2 = data[gender2].fillna(0).sum()

            data

            st.write("--" * 100)

            st.subheader('Total infected: {}'.format(t1 + t2))

            col1, col2 = st.columns(2)

            with col1:
                per1 = (t1 / (t1 + t2)) * 100
                st.subheader('Total of {} infected: {} '.format(gender1, t1))
                st.write('* {}%'.format(per1.__round__(3)))

            with col2:
                per2 = (t2 / (t1 + t2)) * 100
                st.subheader('Total of {} infected: {} '.format(gender2, t2))
                st.write('* {}%'.format(per2.__round__(3)))

            # data
            st.spinner()
            with st.spinner(text='Loading charts'):
                time.sleep(3)
                df = {
                    'Data': [per1.__round__(3),
                             per2.__round__(3)],
                    'index': ['Men', 'Women']
                }
                st.subheader('Men % infected since the first day ')
                # fig = px.pie(df, values='Data', names='Data')
                # st.plotly_chart(fig, use_container_width=True)
                df = DataFrame(df)

                df.plot(kind="bar", color="blue")
                plt.title("Comparative Men % infected since the first day ")
                plt.savefig('D:\\comparative.png')
                st.pyplot()

                st.caption(
                    'Comparative of Men % infected since the first day ')

                interpretacion = "Esta gráfica muestra el porcentaje de hombres infectados frente al total de personas infectadas desde el primer caso de COVID 19 detectado, actualizado a los datos del ultimo dia que contiene el archivo."
                st.info(interpretacion)

                export_as_pdf = st.button("Export Report")
                pdf_title = '11. Porcentaje de hombres infectados por covid-19 en un País desde el primer caso activo'
                #content = ""

                if export_as_pdf:
                    write_pdf(pdf_title, interpretacion, 'D:\\comparative.png')
                #st.line_chart(data[[gender1, gender2]])
                #st.success('Done')
    except Exception as e:
        st.write(e)
        st.warning('Error :c')
    pass


# 12. Ánalisis Comparativo entres 2 o más paises o continentes.
def covidComparative(data: DataFrame):

    try:

        option = st.multiselect(
            'Select field and variable of comparation [place, variable, group by] : ',
            data.columns)

        if option.__len__() < 3:
            st.warning('Select more fields')
        else:

            place = option[0]
            variable = option[1]

            countries = st.multiselect('Select list of countries: ',
                                       data[place].drop_duplicates())
            st.write(countries)

            elements = []
            #size =
            if countries.__len__() >= 2:

                for i in range(0, countries.__len__()):
                    country = countries[i]
                    temp = data[data[place].isin([country])]
                    val = temp[variable].fillna(0).sum()
                    elements.append(val)

                #generateTendencyGraph(elements, '--', 1000000)

                # Graph
                plotdata = pd.DataFrame({
                    str(variable): elements,
                },
                                        index=[countries])

                #st.bar_chart(plotdata,)
                st.spinner()
                with st.spinner(text="loading..."):
                    time.sleep(3)
                    st.set_option('deprecation.showPyplotGlobalUse', False)
                    plotdata.plot(kind="bar", color="blue")
                    plt.title(
                        "Comparative between two or more countries or continents"
                    )
                    plt.savefig("compcountries.png")
                    st.pyplot()

                    interpretacion = """
                    Esta grafica representa la COMPARACIÓN de "{}" entre los paises seleccionados, para poder determinar qué país cuenta con los numeros más altos
                    para poder sacar conclusiones y poder predecir sobre lo que ocurrirá en el futuro con la pandemia.
                    """.format(variable)
                    st.info(interpretacion)

                    export_as_pdf = st.button("Export Report")
                    pdf_title = '12. Ánalisis Comparativo entres 2 o más paises o continentes.'
                    #content = ""

                    if export_as_pdf:
                        write_pdf(pdf_title, interpretacion,
                                  'compcountries.png')
            else:
                st.warning('Please, select more countries ')

    except Exception as e:

        st.write(e)
        st.warning('Error :(')

    # if option == 'Countries':

    # elif option == 'Continents':

    #     col1, col2 = st.columns(2)

    #     with col1:
    #         st.subheader('Select Continent 1:')
    #         continent1 = st.selectbox('Select continent 1: ', data.columns)

    #     with col2:
    #         st.subheader('Select Continent 2:')
    #         continent2 = st.selectbox('Select continent 2: ', data.columns)


# 13. Muertes promedio por casos confirmados y edad de covid 19 en un País.


# 14. Muertes según regiones de un país - Covid 19.
def deathsByCountryRegions(data: DataFrame):

    option_selected = st.multiselect('Select field to filter [country]: ',
                                     data.columns)

    try:

        if option_selected.__len__() == 0:
            st.warning('Please select a filter: ')

        else:
            flter = option_selected[0]

            country = st.selectbox('Select country: ',
                                   data[flter].drop_duplicates())

            data = data[data[flter].isin([country])]

            filters = st.multiselect(
                'Select variables to filter ex: [ state/department/region, deaths]',
                data.columns)

            if filters.__len__() < 2:
                st.warning('Select more filters')

            else:

                data = data.groupby(filters[0])[filters[1]].sum().fillna(0)

                st.subheader("{} by {} in {}".format(filters[1], filters[0],
                                                     country).capitalize())

                st.spinner()
                with st.spinner(text="loading data..."):
                    time.sleep(3)

                    data.plot(kind="bar", color="blue")
                    plt.title("Muertes según regiones de un país - Covid 19.")
                    plt.savefig("deathsregion.png")
                    st.pyplot()
                    data

                    st.info("""
                    Esta tabla representa el numero de "{}" por cada "{}" en {}
                    """.format(filters[1], filters[0], country))

                interpretacion = """
                    Esta grafica representa el numero de "{}" por cada "{}" en {}
                    """.format(filters[1], filters[0], country)

                export_as_pdf = st.button("Export Report")
                pdf_title = "14. Muertes según regiones de un país - Covid 19."
                if export_as_pdf:
                    write_pdf(pdf_title, interpretacion, 'deathsregion.png')

    except Exception as e:
        st.write(e)
        st.warning('Error :c')

    pass


# 15. Tendencia de casos confirmados de Coronavirus en un departamento de un País
def covidCasesByDep(data: DataFrame):

    try:
        # date, state, cases
        data_options = st.multiselect(
            'Select fields [date, region/country, cases, state/province/department]: ',
            data.columns)

        if data_options.__len__() < 4:
            st.warning('Please select more fields. ')
        else:
            st.write(data_options)

            country_option = st.selectbox(
                'Select country', data[data_options[1]].drop_duplicates())

            ## select states
            c = [country_option]
            cs = data[data[data_options[1]].isin(c)]

            # Select department
            province = st.selectbox('Select state/province/department',
                                    cs[data_options[3]].drop_duplicates())

            #st.write(cs)
            # Filter by deparment/state
            dep = cs[data[data_options[3]].isin([province])]

            #st.write(dep)
            #st.write(dep[[data_options[0], data_options[2]]])

            ## convert dates
            dep[data_options[0]] = pd.to_datetime(dep[data_options[0]])
            dep = dep.sort_values(by=data_options[0])

            st.subheader('"{}" data table in {}'.format(
                data_options[2], country_option).capitalize())
            st.write(dep[[data_options[0], data_options[2]]])

            y = dep[data_options[2]].fillna(0)
            hd = "COVID-19 Spread tendency in " + province
            max_val = dep[data_options[2]].max()
            #st.write(max_val)
            generateTendencyGraph(y, hd, max_val)

            interpretacion = """ 
            La gráfica de tendencia de casos confirmados nos muestra el comportamiento que 
            tendrá la variable "{}" para el rango de días que contiene el dataset. Es importante resaltar
            la importancia que tiene la pendiente, ya que si la pendiente es positiva, la tendencia será de aumento
            de casos confirmados, por lo cual deben las autoridades tomar medidas para frentar ese incrementos, pero
            si la pendiente es negativa, esto es un indicativo que la tendencia en los proximos días será de que habrá
            una disminución notable de los casos.
            """.format(data_options[0])
            st.info(interpretacion)

            export_as_pdf = st.button("Export Report")
            pdf_title = '15. Tendencia de casos confirmados de Coronavirus en un departamento de un País'
            #content = ""

            if export_as_pdf:
                write_pdf(pdf_title, interpretacion, 'trend.jpg')

    except Exception as e:
        st.write(e)
        st.warning('Select a field')


# 16. Porcentaje de muertes frente al total de casos en un país, región o continente
def percentageDeathsCases(data: DataFrame):

    try:

        options = st.multiselect(
            'Select filter [country, region or continent] and [death, cases]',
            data.columns)

        if options.__len__() >= 3:

            interpretacion = ""
            st.write(options)
            reg = options[0]
            var1 = options[1]
            var2 = options[2]

            place = st.selectbox('Select place: ', data[reg].drop_duplicates())

            data = data[data[reg].isin([place])]

            # total deaths
            t1 = data[var1].fillna(0).sum()
            t2 = data[var2].fillna(0).sum()
            st.write('* total deaths in {}: {}'.format(place, t1).capitalize())
            interpretacion += '* total deaths in {}: {} \n'.format(
                place, t1).capitalize()
            st.write('* total cases in {}: {}'.format(place, t2).capitalize())
            interpretacion += '* total cases in {}: {} \n'.format(
                place, t2).capitalize()
            perc1 = (t1 / t2) * 100
            perc2 = 100 - perc1
            # st.subheader('Percentege {}, {}'.format(perc1.__round__(2),
            #                                         perc2.__round__(2)))

            interpretacion += '* total percentage deaths in {}: {}% \n'.format(
                place, perc1.__round__(3)).capitalize()
            interpretacion += '* total percentage cases in {}: {}% \n'.format(
                place, perc2.__round__(3)).capitalize()

            df = DataFrame({'Total': [perc1.__round__(3),
                                      perc2.__round__(3)]},
                           index=[var1, var2])

            interpretacion += """\n\n
            El porcemtake de {} es del {} %, mientras que el porcentaje de {} es del {} %
            """.format(var1, perc1.__round__(3), var2, perc2.__round__(3))

            st.spinner()

            with st.spinner(text='Loading data...'):
                time.sleep(5)
                st.subheader('Death percentages in {}'.format(place))
                df

                st.subheader('Deaths percentage vs Positive cases')

                fig = px.pie(df, values='Total', names='Total')
                st.plotly_chart(fig, use_container_width=True)

                st.info(
                    "Esta grafica muestra el porcentaje de muertes que hay entre casos activos en {}"
                    .format(place))

                st.line_chart(data[[var1, var2]])
                st.info("""
                Para poder entender la grafica anterior con mas claridad, la grafica lineal muestra 
                el comportamiento de muertes que hay entre casos activos en {}
                """.format(place))

                export_as_pdf = st.button("Export Report")
                pdf_title = '16. Porcentaje de muertes frente al total de casos en un país, región o continente'
                #content = ""

                if export_as_pdf:
                    write_pdf(
                        pdf_title, interpretacion,
                        'https://htmlcolorcodes.com/assets/images/colors/white-color-solid-background-1920x1080.png'
                    )
        else:
            st.warning('Select more variables')

    except Exception as e:
        st.write(e)
        st.write('Error :C')


# 17. Tasa de comportamiento de casos activos en relación al número de muertes en un continente
def performoranceRateCasesDeaths(data: DataFrame):

    options = st.multiselect('Select filters: [place, date ,cases, deaths]',
                             data.columns)
    st.write(options)

    try:

        if options.__len__() < 4:
            st.warning('Please select more fields')

        else:
            place = options[0]
            date_ = options[1]
            cases = options[2]
            deaths = options[3]
            continent = st.selectbox('Select continent: ',
                                     data[place].drop_duplicates())

            flt = data[data[place].isin([continent])]

            flt[date_] = pd.to_datetime(flt[date_])
            #flt[[date_, cases, deaths]]

            flt[[date_, deaths]]
            st.line_chart(flt[[cases, deaths]].fillna(0))
            st.caption('Cases / Deaths performance during the pandemic. ')
            interpretacion = """
            Esta grafica muestra el comportamiento de las muertes y los casos confirmados a lo largo de la pandemia en {}.
            Este comportamiento suele variar dependiendo cada región del mundo estudiada, en los países más poblados como por ejemplo
            los países que pertenecen a Asia, el numero de muertes y de casos es demasiado grande, que a comparación de otras regiones
            menos pobladas del mundo.
            """.format(continent)

            st.info(interpretacion)

            export_as_pdf = st.button("Export Report")
            pdf_title = '17. Tasa de comportamiento de casos activos en relación al número de muertes en un continente'
            #content = ""

            if export_as_pdf:
                write_pdf(
                    pdf_title, interpretacion,
                    'https://htmlcolorcodes.com/assets/images/colors/white-color-solid-background-1920x1080.png'
                )
        pass
    except Exception as e:

        st.write(e)
        st.warning('Error :(')

    pass


# 18. Comportamiento y clasificación de personas infectadas por COVID-19 por municipio en un País.
def classificationInfectedPeopleByState(data: DataFrame):

    select_field = st.selectbox('Select a field to filter: ', data.columns)

    try:

        country = st.selectbox('Select country',
                               data[select_field].drop_duplicates())

        #data = data[data[select_field].isin([country])]

        country_filter = st.selectbox(
            'Filter country by region/state/province/deparment', data.columns)

        data = data[data[select_field].isin([country])]

        state = st.selectbox('Select state/province/department: ',
                             data[country_filter].drop_duplicates())

        data = data[data[country_filter].isin([state])]

        #data
        filters = st.multiselect('Select variables to filter: ', data.columns)

        if filters.__len__() < 2:
            st.warning('Please select more variables to filter')

        else:
            var1 = filters[0]
            var2 = filters[1]

            has_date = st.checkbox('This file has "date" column ')

            if has_date:

                date_field = st.selectbox('Select date type field',
                                          data.columns)

                data[date_field] = pd.to_datetime(data[date_field])
                data = data.sort_values(by=date_field).reset_index()

                st.subheader('Data classifications')
                st.dataframe(data[[date_field, var1, var2]])

                ## AQUI ME QUEDE xd
                ## columns
                col1, col2 = st.columns(2)

                st.spinner()
                with st.spinner(text="Loading..."):
                    time.sleep(3)
                interpretacion = ''
                with col1:
                    st.set_option('deprecation.showPyplotGlobalUse', False)
                    generateTendencyGraph(
                        data[var1].fillna(0),
                        '{} trend graphic in {}'.format(var1, state),
                        data[var1].fillna(0).max())

                    interpretacion += """\n\n
                    Esta grafica contiene la tendencia que tiene la variable "{}" para el municipio de {}
                    con el paso del tiempo, es importante analizar la pendiente que se genera
                    para saber si irá incrementando o decrementanto el valor de la variable estudiada
                    """.format(var1, state)
                    st.info(interpretacion)

                with col2:
                    st.set_option('deprecation.showPyplotGlobalUse', False)
                    generateTendencyGraph(
                        data[var2].fillna(0),
                        '{} trend graphic in {}'.format(var2, state),
                        data[var2].fillna(0).max())

                    interpretacion += """\n\n
                    Esta grafica contiene la tendencia que tiene la variable "{}" para el municipio de {}
                    con el paso del tiempo, es importante analizar la pendiente que se genera
                    para saber si irá incrementando o decrementanto el valor de la variable estudiada
                    """.format(var2, state)
                    st.info(interpretacion)

                export_as_pdf = st.button("Export Report")
                pdf_title = '18. Comportamiento y clasificación de personas infectadas por COVID-19 por municipio en un País.'
                #content = ""

                if export_as_pdf:
                    write_pdf(
                        pdf_title, interpretacion,
                        'https://htmlcolorcodes.com/assets/images/colors/white-color-solid-background-1920x1080.png'
                    )
            else:
                pass

        #data[[var1, var2]]

    except Exception as e:
        st.write(e)
        st.warning('Error :c')

    pass


# 19. Predicción de muertes en el último día del primer año de infecciones en un país.
def deathsPredictionOnFirstYear(data: DataFrame):

    option = st.selectbox("Select a field to filter [country]: ", data.columns)

    country = st.selectbox("Select country", data[option].drop_duplicates())

    data = data[data[option].isin([country])]

    year = st.selectbox("Select year: ", [2020, 2021])

    date_ = st.selectbox("Select field to filter [date]: ", data.columns)

    fld = st.selectbox("Select field to analyze []: ", data.columns)

    data[date_] = pd.to_datetime(data[date_])
    data = data.sort_values(by=date_)
    include = data[data[date_].dt.year == year]
    include

    n_days = include[date_].max() - include[date_].min()

    st.write(n_days.days)
    st.write(type(n_days))
    #st.write((n_days / np.timedelta64(1, 'D')).astype(int))
    try:

        grade = st.slider('Select a polynomical grade: ', 1, 5)

        generatePredictionGraph(include[fld].fillna(0), grade, n_days.days,
                                include[fld].fillna(0).max())

        global prediccion
        interpretacion = """
        Esta grafica muestra la predicción del numero de muertes para el ultimo dia del primer año de contagios en {},
        con una grafica polinomial de grado {}, y la predicción de muertes para el ultimo dia del primer año de contagios será de {}.
        """.format(country, grade, int(prediccion))

        st.info(interpretacion)

        export_as_pdf = st.button("Export Report")
        pdf_title = '19. Predicción de muertes en el último día del primer año de infecciones en un país.'
        #content = ""

        if export_as_pdf:
            write_pdf(pdf_title, interpretacion, 'prediction.png')

        #st.write(data.loc[data.index.min()-1])
        # st.write(include.index.min())

        pass
    except Exception as e:
        st.write(e)
        st.warning('Error :c ')

    pass


# 20. Tasa de crecimiento de casos de COVID-19 en relación con nuevos casos diarios y tasa de muerte por COVID-19 **
def growthRateCasesAndDeathRate(data: DataFrame):

    try:

        selected_options = st.multiselect(
            'Select variables to analize: [positives, deaths]', data.columns)

        if selected_options.__len__() < 2:
            st.warning('Please select more fields')
        else:
            var1 = selected_options[0]
            var2 = selected_options[1]

            has_date = st.checkbox('This file has Date type field ')

            if has_date:
                date_field = st.selectbox('Select date type field',
                                          data.columns)

                data[date_field] = pd.to_datetime(data[date_field])
                data = data.sort_values(by=date_field).reset_index()
                st.dataframe(data[[date_field, var1, var2]])

                # Tasa de crecimiento de casos de COVID-19 en relación con nuevos casos diarios
                # tasa de muerte por COVID-19
            else:
                pass

    except Exception as e:
        st.write(e)
        st.warning('Error :(')


# 21. Predicciones de casos y muertes en todo el mundo
def deathGlobalPrediction(data: DataFrame):

    try:

        date_ = st.selectbox('Select date field ', data.columns)
        field = st.selectbox('Select variable to study', data.columns)

        n_days = st.slider('Select number of days to predict ', 1, 1000)
        grade = st.slider('Select polynomial grade: ', 1, 10)

        data[date_] = pd.to_datetime(data[date_])
        data = data.sort_values(by=date_)
        st.write(data)

        generatePredictionGraph(data[field].fillna(0), grade, n_days,
                                data[field].fillna(0).max())

        #content = ""
        global prediccion
        interpretacion = """
        La gráfica muestra la predicción de casos asi como tambien la predicción de muertes en el mundo, por causa de la 
        pandemia del COVID-19, representado con un polionomio {} y a un plazo de {} días.
        La predicicón de {} en todo el mundo será de {} personas. 
        """.format(grade, n_days, field, int(prediccion))

        st.info(interpretacion)

        export_as_pdf = st.button("Export Report")
        pdf_title = '21. Predicciones de casos y muertes en todo el mundo '
        if export_as_pdf:
            write_pdf(pdf_title, interpretacion, 'prediction.png')

        pass
    except Exception as e:
        st.write(e)
        st.warning('Error :c')


# 22. Tasa de mortalidad por coronavirus (COVID-19) en un país. ***
def deathsRateByCountry(data: DataFrame):

    option = st.selectbox('Select a field to filter [ex: country]',
                          data.columns)

    country = st.selectbox('Select country: ', data[option].drop_duplicates())

    numeric_vars = st.multiselect(
        'Select numeric varabales to analyze [ex: deaths, positives]: ',
        data.columns)

    if numeric_vars.__len__() == 2:

        deaths = numeric_vars[0]
        positives = numeric_vars[1]

        data = data[data[option].isin([country])]

        flt = data[[deaths, positives]].fillna(0).reset_index()

        st.subheader("Deaths / Positive people in {} ".format(country))
        st.write(flt)

        deathrate = []

        for i in range(0, flt.__len__()):
            # death_rate = (n1/n2) * 100
            n1 = flt[deaths][i]  # deaths
            n2 = flt[positives][i]  # positives

            if n1 == 0 or n2 == 0:
                deathrate.append(0)
            else:
                # st.write('xd')
                death_rate = (n1 / n2) * 100
                deathrate.append(death_rate)

        #xd = pd.DataFrame(deathrate)
        #st.head
        generateTendencyGraph(deathrate, "Death rate by country", 20)
        interpretacion = """
        La grafica de arriba representa el porcentaje de personas fallecidas que hay por el numero de personas 
        que contraen COVID-19 en {}, una pendiente decreeciente indica que la tasa de mortalidad tiende a bajar con el paso del tiempo
        , pero una pendiente positiva indica que la tasa de moratalidad tiene una tendencia a ser más alta en el futuro. Lo cual son
        datos muy importantes para las autoridades, quienes tienen que tomar acciones para evitar 
        que la tasa de mortalidad suba.
        """.format(country)
        st.info(interpretacion)

        st.write('--' * 100)

        df = {'Death Rate': deathrate}

        st.subheader(
            'Grafica lineal de la tasa de mortalidad en {}'.format(country))
        st.line_chart(df)
        st.caption('Death Linear rate graph')
        #de(xd, 3, 20, 1000)

        #st.write(deathrate)
        export_as_pdf = st.button("Export Report")
        pdf_title = '# 22. Tasa de mortalidad por coronavirus (COVID-19) en un país. '
        if export_as_pdf:
            write_pdf(pdf_title, interpretacion, 'trend.jpg')

    pass


# 23. Factores de muerte por COVID-19 en un país.
def covidDeathFactors(data: DataFrame):

    try:

        select_option = st.selectbox('Select field to filter [country]: ',
                                     data.columns)

        country = st.selectbox('Select country:',
                               data[select_option].drop_duplicates())

        data = data[data[select_option].isin([country])]

        select_factors = st.multiselect('Select death factors ', data.columns)

        if select_factors == 0:
            st.write('Please choose more fields')

        else:
            st.write(select_factors)

            elements = []

            for i in range(0, select_factors.__len__()):

                factor = select_factors[i]
                val = data[factor].fillna(0).sum()
                elements.append(val)
                #tot_sum

            plotdata = pd.DataFrame({
                'Deaths': elements,
            },
                                    index=[select_factors])

            st.spinner()
            with st.spinner(text="loading..."):
                time.sleep(3)
                st.set_option('deprecation.showPyplotGlobalUse', False)
                plotdata.plot(kind="bar", color="blue")
                plt.title("Death factors by COUNTRY")
                plt.savefig("deathfactors.png")
                st.pyplot()

                interpretacion = """
                Esta grafica representa la COMPARACIÓN de las causas de muerte de las personas con COVID-19, lo cual es un indicador
                para saber como tiene que estar el estado de salud de las personas de acuerdo a su edad, por lo regular las personas
                con sobrepeso y diabeticas son las que tienen a recibir de peor forma el coronavirus
                """
                st.info(interpretacion)

                export_as_pdf = st.button("Export Report")
                pdf_title = '23. Factores de muerte por COVID-19 en un país.'
                #content = ""

                if export_as_pdf:
                    write_pdf(pdf_title, interpretacion, 'deathfactors.png')

    except Exception as e:
        st.write(e)
        st.warning('Error :c')


# 24. Comparación entre el número de casos detectados y el número de pruebas de un país.
def covidCasesTestComparation(data: DataFrame):

    try:
        options = st.multiselect(
            'Select fields to filter [Country, cases, tests]', data.columns)

        if options.__len__() < 3:
            st.warning('Select more fields.')

        else:
            place = options[0]
            var1 = options[1]
            var2 = options[2]

            country = st.selectbox('Select country',
                                   data[place].drop_duplicates())

            flt = data[data[place].isin([country])]

            val1 = flt[var1].sum()
            val2 = flt[var2].sum()

            #st.write(val1.sum())
            #st.write(val2.sum())
            st.subheader(
                "Comparación entre el numero de casos y numero de pruebas")
            plotdata = pd.DataFrame({'Data': [val1, val2]}, index=[var1, var2])
            plotdata.plot(kind="bar", color="blue")
            plt.title(
                "Comparación entre el numero de casos y numero de pruebas")
            plt.savefig("compcasestests.png")
            st.pyplot()
            #st.bar_chart(plotdata)

            export_as_pdf = st.button("Export Report")
            pdf_title = ' 24. Comparación entre el número de casos detectados y el número de pruebas de un país.'
            #content = ""

            if export_as_pdf:
                write_pdf(
                    pdf_title,
                    "La grafica representa la comparacion entre {}  y {} lo cual es muy util para saber la respuesta de las personas hacia los casos, es decir si hay muchas personas que se hicieron las pruebas por la cantidad de casos detectados.  "
                    .format(var1, var2), 'compcasestests.png')

    except Exception as e:

        st.write(e)
        st.warning('Error :(')


# 25. Predicción de casos confirmados por día
def covidCasesPredictionByDay(data: DataFrame):

    try:

        field = st.multiselect('Select filter, and order by: ', data.columns)

        if field.__len__() < 2:
            st.warning('Please select more fields')
        else:

            flter = field[0]
            order_by = field[1]

            st.write('Filtering by: ', field)

            data[order_by] = pd.to_datetime(data[order_by])

            st.write(data)

            data = data.groupby(order_by)[flter].sum()

            st.subheader('{} field analysis'.format(flter).capitalize())
            st.write(data)
            n_days = st.slider('Select number of days to predict ', 10, 1000)
            grade = st.slider('Select polynomial grade ', 1, 10)
            #y = data[flter]

            st.spinner()
            with st.spinner(text="Generating prediction..."):
                time.sleep(3)
                generatePredictionGraph(data.fillna(0), grade, n_days, data.fillna(0).max())

            global prediccion
            prediccion = """
            La predicción realizada para un periodo de {} días con un polinomio de grado {} brindará un valor proximado en cuanto al numero de "{}".
            Un polomio de grado mayor nos brinda mayor precision que una grafica lineal.
            La predicción de casos confirmados por día es de {} casos para la cantidad de dias seleccionada.
            """.format(n_days, grade, flter, int(prediccion))

            st.info(prediccion)

            export_as_pdf = st.button("Export Report")
            pdf_title = '25. Predicción de casos confirmados por día'
            #content = ""

            if export_as_pdf:
                write_pdf(pdf_title, prediccion, 'prediction.png')

    except Exception as e:

        st.write(e)
        st.warning('Error :(')


# ===================== END METHODS =====================

# Sidebar option tuple
sid_opt_tuple = ('COVID Cases', 'COVID Deaths', 'Vaccines')

#  **** OPTION TUPLES ****
# Covid deaths
covid_deaths_tuple = (
    'Análisis del número de muertes por coronavirus en un País.',
    'Predicción de mortalidad por COVID en un Departamento.',
    'Predicción de mortalidad por COVID en un País',
    'Porcentaje de muertes frente al total de casos en un país, región o continente',
    'Tasa de mortalidad por coronavirus (COVID-19) en un país.',
    'Muertes según regiones de un país - Covid 19.',
    'Predicción de muertes en el último día del primer año de infecciones en un país.',
    'Predicciones de casos y muertes en todo el mundo',
    'Factores de muerte por COVID-19 en un país')
# Covid Cases
covid_cases_tuple = (
    'Tendencia de la infección por Covid-19 en un País.',
    'Indice de Progresión de la pandemia.',
    'Predicción de Infectados en un País.',
    'Ánalisis Comparativo entres 2 o más paises o continentes.',
    'Tendencia del número de infectados por día de un País',
    'Tendencia de casos confirmados de Coronavirus en un departamento de un País',
    'Comparación entre el número de casos detectados y el número de pruebas de un país.',
    'Predicción de casos confirmados por día',
    'Tasa de comportamiento de casos activos en relación al número de muertes en un continente',
    'Predicción de casos de un país para un año.',
    'Porcentaje de hombres infectados por covid-19 en un País desde el primer caso activo',
    'Tasa de crecimiento de casos de COVID-19 en relación con nuevos casos diarios y tasa de muerte por COVID-19',
    'Comportamiento y clasificación de personas infectadas por COVID-19 por municipio en un País.'
)

# Covid Vaccines
covid_vaccines_tuple = ('Tendencia de la vacunación de en un País.',
                        'Ánalisis Comparativo de Vacunación entre 2 paises.')

# Main
st.sidebar.write("""
    # PROYECTO 2
    *Juan Antonio Solares Samayoa* - 201800496
""")

# add sidebar
app_sidebar = st.sidebar.title("MENU")
st.sidebar.header("Load a file: ")

# add selectbox to the sidebar
sidebar_selectbox = st.sidebar.selectbox('Select Type...', sid_opt_tuple)

# select type of file
select_extension = st.sidebar.selectbox('Select a tye of file',
                                        ('json', 'csv', 'xlsx'))

# file uploader
upload_file = st.sidebar.file_uploader("Choose a .csv, .xls or .json file")

# read csv file

# insert image
image = st.sidebar.image(
    'https://peterborough.ac.uk/wp-content/uploads/2020/03/NHS-covid-banner-1.png'
)

image2 = st.image(
    'data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2wCEAAkGBxISEhUTEBIVFhUVGBUVFhUWFRUVFRYVFhcWFxUVFhUYHSggGBolHRUVITEhJSkrLi4uFx8zODMtNygtLisBCgoKDg0OGhAQGy0mICUtLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLS0tLf/AABEIAGoB2wMBIgACEQEDEQH/xAAcAAABBQEBAQAAAAAAAAAAAAAFAgMEBgcBAAj/xABHEAABAwIDBAcDCQYFAgcAAAABAAIRAwQSITEFQVFhBgcTInGBkTKhsSM0QnKSssHR8BQzUnOC4SRTYqLxF2MWZJOjw9Li/8QAGwEAAgMBAQEAAAAAAAAAAAAAAgMAAQQFBgf/xAA3EQABBAAFAQUHAgUFAQAAAAABAAIDEQQSITFBUQUTImGBMnGRobHR8MHhBiM0QnIUM1Ji8ST/2gAMAwEAAhEDEQA/AKk8KHWTzrwcCo9So0713ZJGnYrzMbXDcKPUKYLk7UCZhJWtuyUHKfs20qVninRY57zo1gLj4wN3Ne6P7Hq3ddlCiO88xJ0a0Zue7kBn7t6+jui3Rm3sKQp0G94xjqEd+oeJO4cBoEmSYR+9Pjg7w+Sy3ZnVffPANTsqU7nPLnDyYCPepd51Z3NJpf21Ahoklznsy8S2FsagbX2c2vTNNxc0Egy2JkZjXVZ3Y6cAltXwPzVamdn4dxAfdcnlYLdWD6YBO/LjnwS7FwBLS0d/IOOWEH6UeY9Fq+0ei5ZQLbc4t9RjwD2vDllnA5zM5rPqVi2m8zkZIDTkWxr5pL+2Kw7o8U056sZdjzoRtl0u9wObo64uxLxTZcG4d3dHNqWiqNg75tarYkajcQqGygM6roA5j4hTqjdXOpSZ0nMidU68d3vxG/hyQ6ttIB+JoccogmB9mFzY5cb2jIX6ktvUeFo28IcC32q1Lr0DSF15IsB2VGGaNDq0NOed/EWlrry8AAbuBUupbiMJpgtcc8Jww0aE8T4JmhsljH4wCYzDcteMqOdtO/yxG/MyiVWHNBb3gS0+1g0I3jhw5Qs2IZjsK0NlJaH2DTrB63roa8xYWjCydm45xfC0PdHRFtojkV4bIuzsaJ22CBbP6M3W0H1nUQzExwxtc8MLcU4cjuyOfJFKXVftEfQpf+qPyRvo1dxc4mvdSx/IvfhAkSCCA7I5wA7dJWmbMshRZgD3OGIkYiCROcSOc+q7GF7Se8AMaBQAI6GhQGtkVyuBjOyWREmR5cSSQeupsngHyHyXz/c2b6NR9KoIexxa4cwdx3hJhXTrU2b2d02sBlWbn9enDT/tLPeqfC9NBJ3kYd1XlMRH3cjm9D+fJEthdHri8x9g0HBhxFzsI70wBxORUmp1ZX+KWspQd3aDI8slpPV1s7sbJhI71UmofA5M/wBoB80V6RbRFtbVa29rThne85MH2iFzZMdJ3pYyt6C6sGCjbEHyXtZ1WUUeru9c3SmQZH7wRkYO7iFE2t0OubSl2lVrAwEDuvBMnTJa10Ln9htsRl3Ztk6yd5QvrR+YO/mU/iVceNldMGGt6286VTYKJsLni9r38rVBHVxdVm06mBpENe35QDUAiR6L1r1b38uc+nSknLvg90CBu8fVbBsb5vR/lU/uNU1Zzj5c3HwT29nxZa1+Kw13VztBlQubSY5rmw4NqMkFuhAJHMIBW2PdW9WoLi3fTa4gtc5vdLtDDhkco0O5fSKbrUWvaWvaHNORaQCCOBB1RM7QeCMwFXemip/ZzCCGk3Va6/ZfOFSiHCHCQlwr5066GtoA3FsPk579PXs5+k3/AE8t3ho51X2VtW7YVaTH1KZY5peMUNcCMgcsi3XmukcUzujK0WPn6rkjCSd6IXGvp6Kl2Gy69cxQovqc2tJA8XaDzKXtDodtSIp2TyeOOl+L1v7GgCAAANwyCUue7tKQ+yAPz0XUi7MjbRcSfp+eq+M7ltQvcKkhzXFpHBwMEeoRzo/0Rurs/IUXuG8gZDkXGGjzKufQ3oS28v65qz2dOrUfUIynFUdhptO4mDJ4AxxW7WtqykwMpMaxjRAa0QAOQCqWbJQ3KewGT2dGrAqfVPfgT2AJ4GrTn3GPehW0+jFa1IFxbupzkC4S0ngHglpPgV9KdszFhxNxfwyJ9NUm7tWVWFlVoexwgtcJBHglsxpB1AQyYPMNHG18wi3HAeiW2krb036MfsVeGSaVSXUycyI9phO8iRnwIVeFNdWPK9ocNlxJS5ji124UYUkoUVMFNNXFYMgHVMygJWYk0Ez2K92Sl0wHAEJXZq8oVZioPZLhpqcaaSaarKFMxUA00xWtQUScxNupoSxMbIQgNa1hRqlurBUpKJVt0lzFsjxKCU3PpmWGFZdkbZbU7r8nIXVtlArUC0yMiqa50fuTnsjnFO36q9QuwgWw9szFOpruKsELU1wcLC5EsLo3ZXJEL0JcL0IkukmF6EqF6FSlJMLsJcLkKKUkwvQlwuQopSTC5CchchRVSqj2plwRapbqLVt1hkhJXUZIENcTuXm1DvCkPopPZLPlcNk8PFLYuozZIFKtdlvee7sWHgxoDnx4uIH9C1GtUDWlzjAaCSToABJKp/VA0DZdGP4q0+Pav/CEY6auIsLkt17J/pHe90rE8lz9V046bGK6X+qybpP0vr3lR2F7mUQSGU2ktlu4vj2nHnkFA2Ttq4tnB1Gq5sfRkljuTmnIoXSgiQpVtQc72RMCSu2GxRx+Kg3zqvVefBmmk8Nlx2qyfRaJtXprWqUmsotFJ7wMT8UuEgThBGRmc8/XMBCP0dfVV/ZrJqNjceKOsaRMmc5HIcF47t6EwubGXbAnajqSOL2aBvWm12vf/wANzNnY+VsdWQ3ex4Wg3rW5J2uj0qkI2jWLnYDo07t87woTgpV2PlHKO4L2ODibFh2MaK0HxI1+a8H2hM+fFSPebOYj0BofAaJh4UzYl0Wvwbne50ZKJUC7s5k1WR/EPzPwWftGNsmHka7aifgLHzTOy5nw4yJzN8wHvBIBHzVkZj7s4dO9EnOMsPJXP/F3tu0NDA1sZ4iHVHMEeWfGM1Sy8tLzUc0Myw7oG/ETzVr6uNrfL3Fq46YarPstbUH3D5leKwMJlc5h9nLZrY6jTbjN+i+jdpYgQMZIB4s1DNqRo7XQ85b+e6ndMtl1Kuzml4mrRDHnOT3RFSSNe6SfELMtl2Rr1qdFutRzW+APtHyEnyW+VqYc0tcJDgQRxByIWc9X+wDTva5ePm8saTvc8kB32B/vC9fg5+6gc3pt9PrqvC47D97O11e0denX6WFo9KmGtDWiAAABwAyAWddbW1P3Vs0/9x/vawfePkFoziBmdFhPSC//AGm5q1tznd3kxuTPcAfMpeAjzS5un14R9pS5IsvJ+nK1/od8xtv5bfggfW4+NnuP/cp/FHeiHzK3/lt+CFdZ4mxd9en8UuL+pH+X6p0prDE/9f0R7YR/w1D+VS+41VfrUu30rekab3MJqwS1zmkjA/KQVatjfuKX8qn9xqpfXKxxtqOHXtf/AI3qsNXftvqrxP8AsO9yodt0kuqZxMuak86jnDza6QVqHQTpX+2scyoAK1OC6Mg9pyDwN2eRHhxWI07aoeA8/wAleeqq1LL2cRJNN4O4RLTp4gLq4yGN0RdWo1XKwczmShubQ6UtduKLXtcx4Ba4FrgdCCIIXz7/AOJquxb2u1jG1CMdLC9xAIDgWPOEcBpl7S+h1hu2Ogjtq7X2lguG0hRdbzNMvxGpRExDhEFh9Vy8PIGhzTsRr8V2JY8zmu5CPdVHTq82ldVm3PZhjKQc1lNmEBxeBMklxy5rVFlfVj0XOztoXNB1QVCKLDiDSwQ5zTGEk8eK1RLmADvDtQRxusKo9W9uG0a7hq+4qk+RAA+Pqi3Sy8fRs69Snk9rDhPAkhuLymfJVrq42q3HcWzj3u1fUZzBMPA8CAf6leLig2o1zHgOa4FrgdCCIIKKUZZfEOh9EqE54fCeCPXZfORknESS4mcRJLp4zrPNbh0CvX1rKm+qSXDEwuOrgxxAJO8wAJ5IM7qyt8citUDJ9nukgcA8j8Fcdn2bKNNtKk3Cxgho/M7zvnmtOMxMcrQG7rHgcLLC8uf067qt9Z1mH2RfGdJ7HjzOA/e9yyJrFtvTf5jX8B99sLFH12hwadStHZ5/lkeazdpt/nAjkfRKDVW+kjXNqsM5OHvGv4K0Qg/Sdg7Nrv4ajffqtM4thWPDGpB56KXs1hwCd/8AYfgpWFN2Q7jf1vKCXN1Vq1vkyezaYEfSI1Pqjc7LSFrDITwjxao7ajSYBzTO2ajmW5M5w0H3Shmwg9xE8QfzVGTxhqtsVsL75pHCxNFimFi4WIklQXU0y+kiJppp1JUrBQx9FRa9rKMPppl9NCWhNZKQVVbu3LTI1CsPR3a2Mdm/2hom7u1kIBWaab8TciErVhvhbxlxLMp34WhQvQoGxNoCswH6Q1RKFou1y3MLSWlNwvQlwvQrtCkwvQlQvQqUSYXoSoXYUUSIXYSoXIUUQtzEk01IY3ITrAXQxVS0FD6tsor6CNFiYq0EDowUTHLS+pXaINvVtie9Tf2jR/oqATHg4H7QWi3NBtRjmPEtcC1w4hwgj0K+eNgbUqWVwyvSzLcnN0D2H2mH9agFb3sPbNG7pCrQdiadR9Jrt7XDcVyMVCWOzcFdvBzB7Mp3H0WG9IejlbZ9Ute1xpE9yrHdc3dJ0D+I/BSujtNz34KTHVC6IwCY11OgHMrd3NBEESEmlSa0Q0ADgAB8EEswlgdDILB865tFh8O7DYhs8TqI6i9wQeehWebT6G3DKLX0Xh1QCalMAZ7/AJM7yNM9fcq7Ukt7oh3BwiDOYIOhGa2O5uGU2l9Rwa1okuOQAWUbf24yvWfUwhtIQ1piHO/1O4k+4ALmYrCAxB8TG2w9PasgZSN3a1Q3qxyu3gcc4TmOaR1PBG58NA+IH+0Vd8A5eiFXtqXDENQM43+HJV9982Yh2XKPcTKtgg6JupQa72mNPiAUrs7t92Hj7qVuZo2rQjyo6V5aUn9qfw1Hi5O+idlcd71B89NQSNzzvpzVHXbOPuKN7FtRHaTOId3kNPVSK+zKLxmwCP4cvVO9iA0Gn9FpDW4oadIn015lH2j2yzFQd3FbbOt1VcC750vy+aeyv4edg8SZpadlFtq7vk1XAuvuusrA4yS3C2AdZBGbp9yEdH9sGheMudwfLh/od3XD0PuC5fbfa12AUw9sd/PIneBlmFJo0WC3Di1veaHTmYc46A8tPJH2eP8AReKWM/zC1jdRsd75vnUaiuNVXav/AN/hhlbUQc92h3G3puNDpr5Bby1wIkaHRN06LWlzmgAvIc48SGhoJ8mgeSr/AFf7S7azZiMupfJO/p9g/ZI9CrMtr2lji08LlxvEjQ4c6qt9O9o9jaPAPeq/JN/qnEfsh3uWQYVcOsfaXaXIpA92iIP1nQXe7CPVVIrtYKPJEDydfsuDj5e8lI4Gn3W09E/mdD6jUL6yx/gXfWp/FFOifzOh9QIZ1k/Mj9en8VzIv6kf5fqutP8A0rv8f0R7ZHzel/Lp/cCqnWq0m3pQCfldwn6Lla9j/uKX8un90KYlRvySZuhTpY+8jLLq18/W2z6z3YadJ7jwaxx/DJan0E6NOtWuqVo7WoAMIM4GaxO8k6+AVuTdSoGgucQABJJMAAakk6BPmxjpG5aoLNh8AyJ2YmyuXFZrGue8hrWguc46BrRJJ5ABZN1T9IKT7naVxXqspftFWm9gqPa04ZrQ0YjnhaWAod1pdYjLlpsrB2KmTFasNKkH93TO9s6u0OgkTNQsrcNaB6+O9Mw+FL2nNpamKxYiqtVtew7llXa10+k9r2mhSAcxwc0xhnMZK5rKeqFsXFb+WPvhaskYpmR+XoB9E3Byd5HmPJP1WBXNV1O4fUpuLXNe4hwyIMlXex6wqjO5d25LhEub3CeE03aeqg9Fej/7Td1H1B8lSeSeDn4iWt8N58uK0Lauw7e5/f0g4jR2YcPBwzjktmJmisMe29N+R7lz8LBPTnxuqzsdj5qpXnWQ3Cext3Txe4ADybM+oVt6Pmqbem64M1XAvdlEFxLg2N0AgeSgWPRCzpPD20yXAyMTi4A7iAcvVWFYpnxEVGK8zut+HZNeaV19ANlVOse4DbMs31HMaPI4z933rB9rYm3LSfZcBHiNQtC6X9Jad7XLaLpp0C5gO5zph7hykQDyneqxtOy7WmWjJw7zTwcNF0sPEWwjrdrl4qYHEHpVfv8AFSIQfpQPkD9Zv4p3ZW0ZZFVpa4ZZg5xwUbbFU1KeEgag5btcloeQ5hroskQySC+Cu17sttJb7R7g8Xan0JU7ZFuGUmgcEKYyWBrhkDi89FOt7xzRAggcVBvZ6BR9ZaHUlMdJ6ndZSGr3D7IRHZ1sGMGWvw3Kv9oa1yXu9luQ5Kzm5pge20DxVMNku9Ph+6KQFrWs9T6/slwmhXYThDhP6yQyvtovJZbNLjp2hyaOYGpK5sXZD6Zl5556kog+zpt1/N0Biyi3aHpz+yMYUl1NPwvYUSUobqaafTU5zEy9iiukNqMQbbFt9IKx1WKBeUpaRyQuFik2F+R4Krmx700aoO45FX+m4OAI0Oazi6pq19Eb/GzA45tS43V4VsxseYCQeqPwvQlQvQnLmpML0JUL0KKJMLmFOQvQoqTeFdwpcL0KKIRtC8ZRAL9+QA1KXSrtLMeg1M7lF2ba9rNasJLtBua3cAk9IHAUxRZ7TzpwaMyUOYgF3wC1ZASG/E/X4KVb3bH5NT+BQ9j7N7MAlFMCIXWqEgXpsojqIKd2dd17Z/aW9RzHbyNCODgcnDxTmFdDVRaDoVGuLdQrTa9aFwwRWoMqHi1zqZPiO8PgvXXWzWI+TtabTxdUc8egDfiqnUoAqFWorI7Cx9FsGLlP9ym7V6UXV04GvVkDRg7tMeDRv5mTzTP7cbiJA7heOcE5fDVCrg4QT+p3L2zX4DP6KjWMa5riB4dvKxRr0S++lax4DjTqB86Ni/cVYqFcsyafI6Ig67w0+0cPIanwQJlSj2oe9xGEcSMTSJiPpHPdmndsXtu6mezf3xk3DjGe8OMcOK5PasEEuIjaYjZIJc0bjXQm/W9wOq7/AGNisRDhZXiZtNaQ1jjs7QggdOABQJ9ycuekLMLe4If/ABPAAGEuJOR4Jh10+5onA1zcB77YwgsOhz8FVn15y9OAH6ATtKucRzmYzIJiJygH9Smx9nQR0YhRBBs2737nkWOKu9wFnk7VxU1ic5mkEEABvzAJ399jStSp1OjKdFJ7YwkwN0y30SrV7ZDC4OfE5D8sgp+DJdtq4RJGiN9EOk9S2xmm1pxABzXTAImCI8T+grL/ANRK/wDk0/V/5rOrARVIG9pnyIj9c0VhC/DxONuaibiJYxla6h+dVy6rOe9z3GXPcXE83GSmCU8WpmoE3ZZjqrNs7p1cUKTKTGUy1jQ0Eh0wOMOTO2ul9e6pdlVYwNJDpaHAyNNXFVlO0wkCGMOzBotaHYiUtylxratNvgrfa9O7hjGsaylDWholr5hoAE97kn/+oN1/BS+y/wD+yqTWpUKv9ND/AMQrGLn/AOR+X2Vmr9YF3GQpN5hjj8XLJOl3Sy7vahbWuHupg5UwcNPXexsAnmZVi29cdnRceOQWeUhJlLfFG2g0ALbhJJH297iUY2JbycXD47lYqTEO2PThnjmjFJq0sFBYMQ/NIVYOiG3f2Ko9/ZdpjbhjFhjMGZgq2t6yP/K/+7/+Fn1Nqea1Lkw8bzmcNfX7q48VLG3K00PcPsus6Svtq7n0XFri4lwkYTJJwkH2tVeNldZAc0GrRPMscNfqu09Ssp2/smq84qWfESB8VAsb2rbmKzC3mR3T5oJWRvdT2+qbDJIxtxu9NPot0f1hUI7tKqTwOAD1xFVnbvS6vcgsEUqZyLWklzhwc/hyACrFpdMqCWny/Wq7e18DctTkPxKjMLEw2B8dUEmMmeMpPw0QXajf2es2s32X5PHPc/z0Tte4LtT4Aaf3TF8MQdjMyI/44KJbuLWCdW5eI3FOGhN+/wC/3SaztB5Hz6fZTZgSkYhMcVEdWkGTnuCRWuDTbiOuQHIKGQDXhWITtyn6pg+1yXGu3A8c+KZfUxsa7iYTtBjSO8RIyQ3Z0RVQ1S7amGjLXVSA4OBDhM6hM1HlsRpEcl2lVESASd6IUNEtwJ8SIbMqU6bcMYefHkiVN4cJaZCB0jGfFcuNrigSGgOc6O7wO4n1VkhosoQ0vdQ1U3bu0uwp5e27Jv4lCdj3tVxBcTMjjnO4pdjsypXd2tcyTx0aOACsFrs5lPMCT+tyAZi7Mdun3TXFjW5Rqev2TxCae1SISHtTElQqjVCrBEaoUKuFFSqt7SzKa2NdGlWadxMFEb+nmUEuGwZ4JDtDa7EBD2ZTyFp7TIkb12EP6PXPaUGneMkThOtcktymikQvQlwvQpapIhehLhehS1EmF2F1ehUoq5SvuzaTlA+j+Sa2NSNWoa1XU+yOA3AJhjDo4c548lKDoMbjuUGpBPH5admyggc/RHHZZkx45IcdsU+07NveI1I9nwB3qLTsKT6k1J5NJ7o8OCLUbCm0y1gHBXZtQZa5T2FehcqVGt9ogeJhdDgcwZV2hXExcsT5cmqrskJKl0gm0x3fMKKASBDi3wj8VL2p7J5Qff8A8ofTeCIO9ZnnRaf7VJq1Thw9o50ScJwxHHIKHiExvjn8Uy6WmGsAADs28P8AUT/dedlEiPOdf170kOJTGx0uBx8vH4p62EZHfMYjJPMeibtXA5zzncN390QYA0NJ0iSRAE7st05p0Tb1RSGtKRHZ1RrQG9pRbEy0wHR4yN/JTw4ESCCOIzHqm7BuIAAUy0TiDp7USNC0jjv4J59IMbDRAGgC1+5YTr70xs8TVd9X8Qi0IVsr23+DfeT+SLAIjuqO6bcEw5qlFqTgQoSFEFNOsZCewJQaoqKQAlAJQaukKK6VS6c3ENazjmqxZskor0yq4q8cFC2ezNIdq9dWPwQBWGyEQEWoBC7MItbtWhcp26lUwnoSaQT0KlSbhIrUGvaWvAIOoKi7S2o2iQHDXnClULljwC1w9c1SlHdVXaGzalq7HSJNMnTe08P7qQLl1QB5O6AUU2xXBAYCDJzg7hu/XBANpWzg0NYd/uKEDIDW3T7JuYSEB2/Xy80+XfRf4ymnNGefhzUajdOYIrNkfxAzHipDnxDplv0TqOYPNVmBH580WQtP5R9ybrCGgwBM+McEmmWCQ/PLdon3AOBcdYOXBQrciRkSZ03QhOhCNurSvW8Q+mPrs/L4p4UTq7LDu0lcrAt74bkfUJ8VRUGF3ruUa0bHfhRzj7Q25/VR6l00NJIJaCIE5SpVB+JksEcRuUGrTxuwDQkaaQ3U+anViGCAYA0aNTzUYTZJ2UkaKAG+/ok4nRA/4S7TZ4acTszxTFtVIaTlmfNTaT3YQToiblNFKfmbYGgRSzvGMYe0cGgHInSDw9/qmK/SOnBNJlSpzDS1vqfyTdRjXt7wBaDMHkjtO2YBAYAI0hE4G0phaRqNUN2NtQ1gcTMJ9yJOC7SoNb7LQJ4Jbgq4UO+mihVQoVcIjVaoVZqJUgd8xAbtis94zJALtqB4tbcK5Gegtxk9nmFb4VA6IVcNwBxWhQqadAgxbalPmkQvQlwvQiWZNwvQnF5S1EjCuQnIXYUUVCZXMYXEE574iOaet6hAnLlvz4AoO4SS1u/fymSVLa4MZh1HHnyWZsh+C3PiFUOUWpVBkXHUZ/rcus24xgwMBqOkwGaDkTuQ0MDgATkd0qXaBjCIAABzPJNzkpGRrTrqmLi1r13B78huaJgDhKOWTCxsHxUV+16P+bT+0Fxl614ljgRyQhwG31UeXkaigPJTnVk0+sobqyZdUUslBlK5dVJlCHktPLcptaUw+hiEfqUpwWqPTddohznANyBmXbx4c1Lq7PwxhBgCZOgjQA8SYyTmw2d4h24Ky07VrgJAIBB8xmCmRReG0LpC11Kvs2U7tGgtbiLDnMAxhBGQ8MjxKI0bHFSFRveJbk2YBBmWzudme9uPmEVuLPGWGSMDsWW8QQWnkQYTtGi1jQ1oDQNB71oASS8lRq1kx4bMgtiHAw8RuxJraBgKTc3TWjVRm2b6mb+63h9I/kjA5KEDqmNk04BcfpaeA0+JRVqbfSjTQLrCgJ1VHdOwvYV1pSgrtUkYV4NTqSVLUSYSSlFyae5Vaizbb1TFcP8AFOWAUTaOdd/ip1olNHiJXVk0jA8kfsG5IvQahmzx3Qi1AJ65J3KlUwlQuF7WiXEDxSBeU/4x7x8VN1dhM3+zqdYAVWzGmZB9QoFTo1RiGvqt+rUMehlTLzbNGmJLp93xQOv0oc4xSafIfigdlB8X7/f5Jjc5Hhuvl9l6tbNovLAXOAjNxzkgJmtMZzO4RuXadd7z3gZPH812o+SM4jeeKlCtEJvNqkMo4vaAAGuWuWkKLbUsMgZsOk7uSI0nmCXfSmPJJ7NsYjkI0V5AaIUEhFgqC+m8GWmZE5fil0rGWyRmQIkxmpL6Ry7PIHVRb2XubgcJGRbOf90JaBrVpjXl22n5yuWtwZwvzE4J3TwSb5xa0gERPmm30Q5+pa7Nx8RpCVs60DpfUz1OuZI48kFuIyj4plNac5+HmmKBOI9m4ZgnM5gcPFIoVjOIkSM89/IJ+8dTewuADSwgZZS05Qk0qDCS1pkQDO/wCXRsAFNsZSXD5fqp9kwYcbhrKk9qcERkdELZVOEtYyo48dAn9mMqye00O46J7XbALJJHu4n0ROvXwsEBcd0rjWiT4H+y5DS0AIps7ZdNzcThOZjhkjfZ2P6pURaNCLQ1vSpx9m2ef6v7IftHbd0/IN7McBr5lXFljTGjB8fik1dnU3fRjwy/slltisx+SaHgH2R6koNsS5qOAFQzkdeSmVmqXTtWs9keZ1TNdqNJPVCLtuRQG8YrFdNQS9aodk2A0UP2M/DcM8VqCyuzyrs+stVaMktuy04v2m+5cheTkL0K7WNNwuwlwuQpaiRC9CXC7ClqLJHEsJE4Tw4854JdGXZak5znknL3U/ymfErzP3h81kLKK6me23ypVNgnMqRSJg8M1Hp+yfFLrewPFNAAWZws0pdOg3CBhGgnL1XKFqGzhGqkWmbROaeaiDQgN6qN2RTjLYqUxSqaPKhQh9suNoIjXTSU5MJoKbsywa8OnIiII1GqfNCrT0GMcW6+bdfSU/sXR3l+KJLRGaaszTog7bp5yDHzwwlPMtKjvbdhHAZn10HvRELqPP0RWolCyY3MCTxdmfLcPJSIS11CTaFNOZKZfSUpcKpWoYlKDk65IUKlLmNJLkpcCpRIzTjaScanGq1Fk+1GxXf4qRalNbc+c1PFKtELfaK6b9WA+StWz/ZCLUEJ2b7IRegmlcnlMbYs6jxNPMxEb/JV+z2DdVHd89m3eTr5DerpQTyW4ZqtMY8tuuUGtOjNBmbmmo7+Kp3vdoilOi1uTQB4AD4J5cUGmyokndVXabMNZ7RIGRA8Wgwh2HdGco1tH5y76rfghl7u8Sjq22gDvHlSCxstYRpO/ko1faTA7CWzByjVSKXszvxjPzUK4Hyzo4fige4hthNjaC6nJ03kw0Me0HjklOsILXMOc5k6KE95w6ndvRO4/cs8CrAzWXcaqyctBvJpMOuCHkYZOWY3BN31Xsmy32p13Z8UYt6Yw6DQbuSCbR9jzH4qPsMJv9lcZBeBW2h81xlIPJa4NGWLuTBKYrWz2OBxTMAeGgU+0Gn1Ch139HxS5GAC+f8AxOjcS6uETN4Iw4iCJkjiFyncPcMzp70NIzd5I0PZ9ETXF6TI1rK03UizbJkDKEUsdo06cse4CMx56hQ7fT0Ve2w0dochu+COU5GWs8De9kI2Vs2nt+nTHcMnj+QRDZdz2tMPI5ePP3rO30mwO6PQLQ9jD5Gn9UJTJM16LRLFkrVO1GqLWaptVRKqMJJQy4age0W6qwXCBbVRK4vbCB2Ymuz6y1doyWV7L+cM8Vq4SgtmL3b7kmF6EpKVrIm4XoS15UokQvQlLytUv//Z'
)

if upload_file is not None:

    st.write(select_extension)
    st.write(upload_file)
    if select_extension == 'json':
        data = pd.read_json(upload_file)
    elif select_extension == 'csv':
        data = pd.read_csv(upload_file)
    elif select_extension == 'xlsx':
        data = pd.read_excel(upload_file, sheet_name=0)

    # Validate area of analysis
    if sidebar_selectbox == 'COVID Cases':
        st.header('COVID Spread/Cases Reports ')

        select_report = st.selectbox('Select report', covid_cases_tuple)
        data.replace(np.nan, 0)

        if select_extension == 'json':
            st.write(data.head().to_string())
        elif select_extension == 'csv':
            data
        elif select_extension == 'xlsx':
            st.write(data.head().to_string())
            pass

        # Validate option
        if select_report == 'Tendencia de la infección por Covid-19 en un País.':
            covidInfectionTendence(data)

        elif select_report == 'Tendencia de casos confirmados de Coronavirus en un departamento de un País':
            covidCasesByDep(data)

        elif select_report == 'Indice de Progresión de la pandemia.':
            pandemicProgression(data)

        elif select_report == 'Predicción de Infectados en un País.':
            covidInfectedPredictionByCountry(data)

        elif select_report == 'Tendencia del número de infectados por día de un País':
            #select_date = st.selectbox('Select data to parameterize', data.)
            covidInfectedByDay(data)

        elif select_report == 'Ánalisis Comparativo entres 2 o más paises o continentes.':

            covidComparative(data)

        elif select_report == 'Comparación entre el número de casos detectados y el número de pruebas de un país.':

            covidCasesTestComparation(data)

        elif select_report == 'Predicción de casos confirmados por día':

            covidCasesPredictionByDay(data)

        elif select_report == 'Tasa de comportamiento de casos activos en relación al número de muertes en un continente':

            performoranceRateCasesDeaths(data)

        elif select_report == 'Predicción de casos de un país para un año.':

            casesPredictionOneYear(data)

        elif select_report == 'Porcentaje de hombres infectados por covid-19 en un País desde el primer caso activo':

            menPercentageInfected(data)

        elif select_report == 'Tasa de crecimiento de casos de COVID-19 en relación con nuevos casos diarios y tasa de muerte por COVID-19':

            growthRateCasesAndDeathRate(data)

        elif select_report == 'Comportamiento y clasificación de personas infectadas por COVID-19 por municipio en un País.':
            classificationInfectedPeopleByState(data)

        #report_text2 = st.text_input("Content")
        # export_as_pdf = st.button("Export Report")

        # if export_as_pdf:
        #     pdf = FPDF()
        #     pdf.add_page()
        #     pdf.set_font('Arial', 'B', 16)
        #     pdf.cell(40, 10, report_text)
        #     pdf.cell(40, 10, '\n\n\n')

        #     html = create_download_link(pdf.output(dest="S").encode("latin-1"), "test")

        #     st.markdown(html, unsafe_allow_html=True)

    elif sidebar_selectbox == 'COVID Deaths':

        st.header("COVID deaths Reports")
        select_report = st.selectbox('Select report', covid_deaths_tuple)
        data.replace(np.nan, 0)
        st.write(data)
        if select_report == 'Análisis del número de muertes por coronavirus en un País.':
            covidDeathsByCountry(data)

        elif select_report == 'Predicción de mortalidad por COVID en un Departamento.':
            covidDeathsPredictionByDeparment(data)

        elif select_report == 'Predicción de mortalidad por COVID en un País':
            covidDeathPredictionByCountry(data)

        elif select_report == 'Porcentaje de muertes frente al total de casos en un país, región o continente':
            percentageDeathsCases(data)

        elif select_report == 'Tasa de mortalidad por coronavirus (COVID-19) en un país.':
            deathsRateByCountry(data)

        elif select_report == 'Muertes según regiones de un país - Covid 19.':
            deathsByCountryRegions(data)

        elif select_report == 'Predicción de muertes en el último día del primer año de infecciones en un país.':
            deathsPredictionOnFirstYear(data)

        elif select_report == 'Predicciones de casos y muertes en todo el mundo':
            deathGlobalPrediction(data)

        elif select_report == 'Factores de muerte por COVID-19 en un país':
            covidDeathFactors(data)

    elif sidebar_selectbox == 'Vaccines':
        st.header('COVID Vaccines Reports')
        select_report = st.selectbox('Select report', covid_vaccines_tuple)
        data.replace(np.nan, 0)
        st.write(data)
        if select_report == 'Tendencia de la vacunación de en un País.':
            vaccinationTendencyByCountry(data)

        elif select_report == 'Ánalisis Comparativo de Vacunación entre 2 paises.':
            vaccinationComparationByCountries(data)

else:

    st.warning("The file is empty or invalid, please upload a valid file")

# ## Validate sidebar option
